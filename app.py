import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime
from difflib import SequenceMatcher
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


st.set_page_config(
    page_title="AI Travel Agent Pro",
    page_icon="🌍",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700;800;900&display=swap');

* { font-family: 'Inter', sans-serif; }

:root {
    --glass: rgba(255,255,255,0.085);
    --glass-border: rgba(255,255,255,0.18);
    --cyan: #06b6d4;
    --violet: #8b5cf6;
    --pink: #ec4899;
    --green: #22c55e;
}

.stApp {
    background:
        radial-gradient(circle at 10% 15%, rgba(45, 212, 191, 0.28), transparent 28%),
        radial-gradient(circle at 85% 10%, rgba(168, 85, 247, 0.24), transparent 34%),
        radial-gradient(circle at 50% 95%, rgba(236, 72, 153, 0.16), transparent 35%),
        linear-gradient(135deg, #020617 0%, #08111f 48%, #030712 100%);
    color: white;
    overflow-x: hidden;
}

.stApp::before {
    content: "";
    position: fixed;
    inset: 0;
    background-image:
        linear-gradient(rgba(255,255,255,.035) 1px, transparent 1px),
        linear-gradient(90deg, rgba(255,255,255,.035) 1px, transparent 1px);
    background-size: 48px 48px;
    animation: gridMove 18s linear infinite;
    pointer-events: none;
    z-index: 0;
    mask-image: linear-gradient(to bottom, rgba(0,0,0,.8), transparent 90%);
}

.stApp::after {
    content: "";
    position: fixed;
    inset: 0;
    background: radial-gradient(circle, rgba(255,255,255,.10) 1px, transparent 1px);
    background-size: 70px 70px;
    animation: starFloat 26s linear infinite;
    pointer-events: none;
    z-index: 0;
    opacity: .45;
}

@keyframes gridMove {
    from { transform: translateY(0px); }
    to { transform: translateY(48px); }
}

@keyframes starFloat {
    from { transform: translate3d(0,0,0); }
    to { transform: translate3d(-70px,70px,0); }
}

.block-container {
    position: relative;
    z-index: 2;
    padding-top: 2rem;
}

/* Flying 3D airplane */
.plane3d {
    position: fixed;
    top: 16%;
    left: -160px;
    font-size: 74px;
    z-index: 5;
    filter: drop-shadow(0 22px 22px rgba(0,0,0,.65));
    animation: flyPlane 15s linear infinite;
    pointer-events: none;
}

.plane-trail {
    position: fixed;
    top: 21%;
    left: -260px;
    width: 210px;
    height: 3px;
    z-index: 4;
    background: linear-gradient(90deg, transparent, rgba(125,211,252,.65), transparent);
    box-shadow: 0 0 25px rgba(125,211,252,.8);
    animation: flyTrail 15s linear infinite;
    pointer-events: none;
}

@keyframes flyPlane {
    0% { transform: translateX(-180px) translateY(20px) rotate(-12deg) scale(.72); opacity: 0; }
    8% { opacity: 1; }
    42% { transform: translateX(48vw) translateY(-55px) rotate(7deg) scale(1.20); opacity: 1; }
    74% { opacity: 1; }
    100% { transform: translateX(125vw) translateY(35px) rotate(17deg) scale(.82); opacity: 0; }
}

@keyframes flyTrail {
    0% { transform: translateX(-280px) translateY(20px) rotate(-8deg); opacity: 0; }
    8% { opacity: .9; }
    42% { transform: translateX(47vw) translateY(-48px) rotate(6deg); opacity: .9; }
    100% { transform: translateX(124vw) translateY(35px) rotate(15deg); opacity: 0; }
}

/* Floating 3D objects */
.glow-orb-one, .glow-orb-two, .cloud-one, .cloud-two {
    position: fixed;
    pointer-events: none;
    z-index: 1;
}

.glow-orb-one {
    width: 240px;
    height: 240px;
    border-radius: 50%;
    left: 4%;
    bottom: 12%;
    background: radial-gradient(circle, rgba(6,182,212,.36), transparent 68%);
    filter: blur(2px);
    animation: orbOne 9s ease-in-out infinite alternate;
}

.glow-orb-two {
    width: 190px;
    height: 190px;
    border-radius: 50%;
    right: 8%;
    top: 18%;
    background: radial-gradient(circle, rgba(236,72,153,.30), transparent 68%);
    animation: orbTwo 11s ease-in-out infinite alternate;
}

.cloud-one, .cloud-two {
    font-size: 54px;
    opacity: .22;
    filter: blur(.2px) drop-shadow(0 20px 28px rgba(0,0,0,.35));
}

.cloud-one { top: 38%; left: 3%; animation: cloudMoveOne 24s linear infinite; }
.cloud-two { top: 62%; right: 7%; animation: cloudMoveTwo 28s linear infinite; }

@keyframes orbOne {
    from { transform: translate3d(0,0,0) scale(.95); }
    to { transform: translate3d(60vw,-28vh,0) scale(1.25); }
}

@keyframes orbTwo {
    from { transform: translate3d(0,0,0) scale(1); }
    to { transform: translate3d(-46vw,42vh,0) scale(1.35); }
}

@keyframes cloudMoveOne {
    from { transform: translateX(-10vw) translateY(0); }
    to { transform: translateX(110vw) translateY(-25px); }
}

@keyframes cloudMoveTwo {
    from { transform: translateX(15vw) translateY(0); }
    to { transform: translateX(-115vw) translateY(30px); }
}

.hero {
    position: relative;
    overflow: hidden;
    padding: 38px;
    border-radius: 34px;
    background:
        linear-gradient(135deg, rgba(59,130,246,.30), rgba(168,85,247,.20)),
        linear-gradient(180deg, rgba(255,255,255,.12), rgba(255,255,255,.03));
    border: 1px solid rgba(255,255,255,.20);
    box-shadow:
        0 34px 95px rgba(0,0,0,.58),
        inset 0 1px 0 rgba(255,255,255,.28),
        inset 0 -45px 80px rgba(0,0,0,.18);
    transform: perspective(1100px) rotateX(2deg);
    transition: .45s ease;
    margin-bottom: 28px;
}

.hero::before {
    content: "";
    position: absolute;
    top: -80px;
    right: -120px;
    width: 360px;
    height: 360px;
    border-radius: 50%;
    background: radial-gradient(circle, rgba(125,211,252,.45), transparent 68%);
    animation: pulseGlow 4s ease-in-out infinite alternate;
}

.hero::after {
    content: "✈️  🌍  🧭";
    position: absolute;
    right: 34px;
    bottom: 20px;
    font-size: 44px;
    opacity: .25;
    transform: rotate(-8deg);
}

.hero:hover {
    transform: perspective(1100px) rotateX(5deg) rotateY(-4deg) scale(1.012);
    box-shadow: 0 45px 120px rgba(6,182,212,.25), 0 28px 80px rgba(0,0,0,.60);
}

.hero h1 {
    font-size: 52px;
    font-weight: 900;
    margin-bottom: 8px;
    background: linear-gradient(90deg, #ffffff, #bae6fd, #ddd6fe);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    text-shadow: 0 12px 40px rgba(6,182,212,.18);
}

.hero p {
    color: #cbd5e1;
    font-size: 17px;
    max-width: 900px;
}

@keyframes pulseGlow {
    from { transform: scale(.92); opacity: .55; }
    to { transform: scale(1.15); opacity: .95; }
}

.metric3d {
    position: relative;
    overflow: hidden;
    padding: 22px;
    border-radius: 24px;
    background:
        linear-gradient(145deg, rgba(255,255,255,.14), rgba(255,255,255,.045));
    border: 1px solid rgba(255,255,255,.17);
    box-shadow:
        0 22px 55px rgba(0,0,0,.42),
        inset 0 1px 0 rgba(255,255,255,.22);
    transform-style: preserve-3d;
    transition: transform .35s ease, box-shadow .35s ease, border .35s ease;
}

.metric3d::before {
    content: "";
    position: absolute;
    inset: -1px;
    background: linear-gradient(120deg, transparent, rgba(255,255,255,.22), transparent);
    transform: translateX(-120%);
    transition: .65s ease;
}

.metric3d:hover::before { transform: translateX(120%); }

.metric3d:hover {
    transform: perspective(1000px) rotateX(8deg) rotateY(-8deg) translateY(-9px) scale(1.02);
    box-shadow: 0 34px 88px rgba(6,182,212,.28), 0 20px 60px rgba(0,0,0,.55);
    border-color: rgba(125,211,252,.42);
}

.metric3d h3 {
    color: #a5b4fc;
    font-size: 14px;
    margin-bottom: 9px;
    letter-spacing: .3px;
}

.metric3d h2 {
    font-size: 29px;
    font-weight: 900;
    margin: 0;
    color: #f8fafc;
}

/* Streamlit cards, tabs and widgets */
[data-testid="stVerticalBlockBorderWrapper"],
[data-testid="stExpander"],
.stDataFrame,
[data-testid="stPlotlyChart"] {
    border-radius: 24px !important;
    background: rgba(255,255,255,.055) !important;
    border: 1px solid rgba(255,255,255,.12) !important;
    box-shadow: 0 20px 55px rgba(0,0,0,.26) !important;
    backdrop-filter: blur(18px);
}

.stTabs [data-baseweb="tab-list"] {
    gap: 10px;
    background: rgba(255,255,255,.06);
    border: 1px solid rgba(255,255,255,.12);
    padding: 10px;
    border-radius: 22px;
    box-shadow: 0 18px 45px rgba(0,0,0,.28);
}

.stTabs [data-baseweb="tab"] {
    border-radius: 16px;
    color: #cbd5e1;
    font-weight: 800;
    padding: 12px 18px;
    transition: .25s ease;
}

.stTabs [aria-selected="true"] {
    background: linear-gradient(135deg, rgba(6,182,212,.35), rgba(139,92,246,.35));
    color: white !important;
    box-shadow: 0 12px 34px rgba(99,102,241,.30);
}

.stButton button, .stDownloadButton button {
    border-radius: 17px !important;
    font-weight: 900 !important;
    border: 1px solid rgba(255,255,255,.14) !important;
    background: linear-gradient(135deg, #06b6d4, #8b5cf6, #ec4899) !important;
    color: white !important;
    box-shadow: 0 14px 38px rgba(99,102,241,.42), inset 0 1px 0 rgba(255,255,255,.26);
    transition: .25s ease !important;
}

.stButton button:hover, .stDownloadButton button:hover {
    transform: translateY(-3px) scale(1.015);
    box-shadow: 0 20px 55px rgba(6,182,212,.45);
}

input, textarea, [data-baseweb="select"] > div {
    border-radius: 16px !important;
    background: rgba(255,255,255,.08) !important;
    border: 1px solid rgba(255,255,255,.14) !important;
    color: white !important;
    box-shadow: inset 0 1px 0 rgba(255,255,255,.12);
}

[data-testid="stSidebar"] {
    background:
        radial-gradient(circle at top, rgba(6,182,212,.16), transparent 28%),
        linear-gradient(180deg, rgba(2,6,23,.98), rgba(15,23,42,.96));
    border-right: 1px solid rgba(255,255,255,.12);
    box-shadow: 18px 0 55px rgba(0,0,0,.30);
}

[data-testid="stSidebar"] * {
    color: #e5e7eb;
}

/* Alert polish */
[data-testid="stAlert"] {
    border-radius: 18px;
    border: 1px solid rgba(255,255,255,.14);
    box-shadow: 0 16px 40px rgba(0,0,0,.25);
}

/* Reduce motion on small devices */
@media (max-width: 768px) {
    .plane3d, .plane-trail, .cloud-one, .cloud-two { display: none; }
    .hero h1 { font-size: 36px; }
}
</style>
""", unsafe_allow_html=True)


st.markdown("""
<div class="plane3d">✈️</div>
<div class="plane-trail"></div>
<div class="glow-orb-one"></div>
<div class="glow-orb-two"></div>
<div class="cloud-one">☁️</div>
<div class="cloud-two">☁️</div>
""", unsafe_allow_html=True)


def clean_money(value):
    if pd.isna(value):
        return 0.0
    value = str(value)
    value = re.sub(r"[^0-9.]", "", value)
    try:
        return float(value)
    except:
        return 0.0


@st.cache_data
def load_data(file):
    df = pd.read_csv(file)

    df.rename(columns={
        "Traveler gender": "Gender",
        "Traveler age": "Age",
        "Traveler nationality": "Nationality",
        "Accommodation type": "Accommodation",
        "Accommodation cost": "Hotel Cost",
        "Transportation type": "Transport",
        "Transportation cost": "Transport Cost",
        "Duration (days)": "Duration"
    }, inplace=True)

    required = [
        "Destination", "Gender", "Age", "Nationality",
        "Accommodation", "Hotel Cost", "Transport",
        "Transport Cost", "Duration"
    ]

    for col in required:
        if col not in df.columns:
            df[col] = "Unknown"

    df["Hotel Cost"] = df["Hotel Cost"].apply(clean_money)
    df["Transport Cost"] = df["Transport Cost"].apply(clean_money)
    df["Age"] = pd.to_numeric(df["Age"], errors="coerce").fillna(0)
    df["Duration"] = pd.to_numeric(df["Duration"], errors="coerce").fillna(0)

    df["Total Trip Cost"] = df["Hotel Cost"] + df["Transport Cost"]

    for col in ["Destination", "Gender", "Nationality", "Accommodation", "Transport"]:
        df[col] = df[col].fillna("Unknown").astype(str)

    return df


def make_excel_report(df, filtered_df, budget, trip_type):
    wb = Workbook()
    ws = wb.active
    ws.title = "Travel Report"

    title_fill = PatternFill("solid", fgColor="111827")
    header_fill = PatternFill("solid", fgColor="2563EB")
    white_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=18)
    border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = "AI Travel Agent Professional Report"
    ws["A1"].font = title_font
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Generated On"
    ws["B3"] = datetime.now().strftime("%d-%m-%Y %I:%M %p")
    ws["A4"] = "Selected Budget"
    ws["B4"] = f"₹{budget:,.0f}"
    ws["A5"] = "Trip Type"
    ws["B5"] = trip_type.title()

    summary = [
        ["Total Travelers", len(filtered_df)],
        ["Unique Destinations", filtered_df["Destination"].nunique()],
        ["Average Hotel Cost", round(filtered_df["Hotel Cost"].mean(), 2)],
        ["Average Transport Cost", round(filtered_df["Transport Cost"].mean(), 2)],
        ["Average Total Cost", round(filtered_df["Total Trip Cost"].mean(), 2)],
        ["Average Duration", round(filtered_df["Duration"].mean(), 2)]
    ]

    ws["A7"] = "Summary"
    ws["A7"].font = white_font
    ws["A7"].fill = header_fill

    row = 8
    for k, v in summary:
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        row += 1

    row += 2
    ws[f"A{row}"] = "Top Destinations"
    ws[f"A{row}"].font = white_font
    ws[f"A{row}"].fill = header_fill
    row += 1

    top_dest = filtered_df["Destination"].value_counts().head(10).reset_index()
    top_dest.columns = ["Destination", "Trips"]

    for col_num, col_name in enumerate(top_dest.columns, 1):
        cell = ws.cell(row=row, column=col_num, value=col_name)
        cell.font = white_font
        cell.fill = header_fill

    row += 1
    for _, r in top_dest.iterrows():
        ws.cell(row=row, column=1, value=r["Destination"])
        ws.cell(row=row, column=2, value=int(r["Trips"]))
        row += 1

    row += 2
    ws[f"A{row}"] = "Recommended Affordable Trips"
    ws[f"A{row}"].font = white_font
    ws[f"A{row}"].fill = header_fill
    row += 1

    cols = ["Destination", "Accommodation", "Hotel Cost", "Transport", "Transport Cost", "Total Trip Cost"]
    affordable = filtered_df[filtered_df["Total Trip Cost"] <= budget][cols].head(20)

    for col_num, col_name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=col_num, value=col_name)
        cell.font = white_font
        cell.fill = header_fill

    row += 1
    for _, r in affordable.iterrows():
        for col_num, col_name in enumerate(cols, 1):
            ws.cell(row=row, column=col_num, value=r[col_name])
        row += 1

    for col in "ABCDEF":
        ws.column_dimensions[col].width = 24

    for rows in ws.iter_rows():
        for cell in rows:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    file = BytesIO()
    wb.save(file)
    file.seek(0)
    return file.getvalue()


def similarity(a, b):
    return SequenceMatcher(None, str(a).lower(), str(b).lower()).ratio()


def normalize_text(text):
    text = str(text).lower()
    text = re.sub(r"[^a-z0-9\s]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def find_best_column(question, df):
    q = normalize_text(question)

    column_keywords = {
        "Destination": [
            "destination", "place", "city", "country", "location",
            "visited", "visit", "trip place", "tourist place", "popular place"
        ],
        "Gender": [
            "gender", "male", "female", "men", "women", "boy", "girl"
        ],
        "Age": [
            "age", "old", "young", "traveler age", "average age"
        ],
        "Nationality": [
            "nationality", "traveler country", "from where", "people from"
        ],
        "Accommodation": [
            "hotel", "stay", "accommodation", "room", "resort", "hostel"
        ],
        "Hotel Cost": [
            "hotel cost", "hotel price", "stay cost", "room price",
            "accommodation cost"
        ],
        "Transport": [
            "transport", "vehicle", "bus", "flight", "train", "car",
            "cab", "aeroplane", "plane", "ship"
        ],
        "Transport Cost": [
            "transport cost", "travel cost", "ticket price", "fare",
            "bus price", "flight price"
        ],
        "Duration": [
            "duration", "days", "trip days", "how many days", "long trip"
        ],
        "Total Trip Cost": [
            "total cost", "overall cost", "budget", "expense", "cheap",
            "costly", "price", "low budget", "high budget"
        ]
    }

    best_col = None
    best_score = 0

    for col, keywords in column_keywords.items():
        if col not in df.columns:
            continue

        for word in keywords:
            score = similarity(q, word)

            if word in q:
                score += 0.60

            for token in q.split():
                score = max(score, similarity(token, word))

            if score > best_score:
                best_score = score
                best_col = col

    return best_col


def search_dataset_values(df, question):
    q = normalize_text(question)
    matches = []

    object_cols = df.select_dtypes(include=["object"]).columns.tolist()

    for col in object_cols:
        unique_values = df[col].dropna().astype(str).unique()

        for value in unique_values:
            value_clean = normalize_text(value)

            if not value_clean or value_clean == "unknown":
                continue

            score = similarity(q, value_clean)

            if value_clean in q:
                score += 0.70

            for token in q.split():
                score = max(score, similarity(token, value_clean))

            if score >= 0.62:
                matches.append((score, col, value))

    matches = sorted(matches, reverse=True, key=lambda x: x[0])
    return matches[:5]


def format_table(df, cols=None, limit=10):
    if df.empty:
        return "No matching records found."

    if cols:
        available_cols = [c for c in cols if c in df.columns]
        df = df[available_cols]

    return df.head(limit).to_string(index=False)


def answer_question(df, question):
    q = normalize_text(question)

    if df.empty:
        return "No data available after applying filters."

    matched_col = find_best_column(q, df)
    value_matches = search_dataset_values(df, q)

    numeric_cols = ["Age", "Hotel Cost", "Transport Cost", "Duration", "Total Trip Cost"]
    common_cols = [
        "Destination", "Gender", "Nationality", "Accommodation",
        "Hotel Cost", "Transport", "Transport Cost", "Duration", "Total Trip Cost"
    ]

    if value_matches:
        _, col, value = value_matches[0]
        matched_rows = df[df[col].astype(str).str.lower() == str(value).lower()]

        if not matched_rows.empty:
            if any(w in q for w in ["average", "avg", "mean"]):
                result = []
                for ncol in numeric_cols:
                    if ncol in matched_rows.columns:
                        result.append(f"{ncol}: {matched_rows[ncol].mean():,.2f}")
                return f"Data found for '{value}' in {col}.\n\nAverage values:\n" + "\n".join(result)

            if any(w in q for w in ["count", "how many", "number", "total"]):
                return f"'{value}' found in column '{col}'.\nTotal matching records: {len(matched_rows)}"

            return (
                f"Data found for '{value}' in column '{col}'.\n\n"
                + format_table(matched_rows, common_cols, 10)
            )

    if any(w in q for w in ["summary", "overview", "report", "insight", "details"]):
        return f"""
Smart Dataset Summary

Total Travelers: {len(df)}
Unique Destinations: {df['Destination'].nunique()}
Top Destination: {df['Destination'].value_counts().idxmax()}
Most Used Transport: {df['Transport'].value_counts().idxmax()}
Most Common Accommodation: {df['Accommodation'].value_counts().idxmax()}
Average Age: {df['Age'].mean():.1f}
Average Hotel Cost: ₹{df['Hotel Cost'].mean():,.0f}
Average Transport Cost: ₹{df['Transport Cost'].mean():,.0f}
Average Total Trip Cost: ₹{df['Total Trip Cost'].mean():,.0f}
Average Duration: {df['Duration'].mean():.1f} days
"""

    if any(w in q for w in ["male", "men", "boys"]):
        male_df = df[df["Gender"].astype(str).str.lower().str.contains("male", na=False)]
        female_word_present = any(w in q for w in ["female", "women", "girls"])

        if not female_word_present:
            return f"Male travelers found: {len(male_df)}"

    if any(w in q for w in ["female", "women", "girls"]):
        female_df = df[df["Gender"].astype(str).str.lower().str.contains("female", na=False)]
        return f"Female travelers found: {len(female_df)}"

    if any(w in q for w in ["how many", "count", "number", "total"]):
        if matched_col and matched_col in df.columns:
            counts = df[matched_col].value_counts().head(10)
            return f"Top counts from {matched_col}:\n\n{counts.to_string()}"
        return f"Total records found: {len(df)}"

    if any(w in q for w in ["top", "most", "popular", "famous", "maximum", "highest"]):
        if matched_col and matched_col in df.columns:
            if matched_col in numeric_cols:
                top_rows = df.sort_values(matched_col, ascending=False)
                return format_table(top_rows, common_cols, 10)

            counts = df[matched_col].value_counts().head(10)
            return f"Top {matched_col} values:\n\n{counts.to_string()}"

        counts = df["Destination"].value_counts().head(10)
        return f"Top Destinations:\n\n{counts.to_string()}"

    if any(w in q for w in ["average", "avg", "mean"]):
        if matched_col in numeric_cols:
            return f"Average {matched_col}: {df[matched_col].mean():,.2f}"

        result = []
        for col in numeric_cols:
            if col in df.columns:
                result.append(f"{col}: {df[col].mean():,.2f}")
        return "Average values:\n\n" + "\n".join(result)

    if any(w in q for w in ["cheap", "cheapest", "minimum", "lowest", "low budget", "affordable", "less price"]):
        cheap = df.sort_values("Total Trip Cost").head(10)
        return "Cheapest trips:\n\n" + format_table(cheap, common_cols, 10)

    if any(w in q for w in ["expensive", "costly", "maximum price", "highest price", "high budget"]):
        expensive = df.sort_values("Total Trip Cost", ascending=False).head(10)
        return "Most expensive trips:\n\n" + format_table(expensive, common_cols, 10)

    if matched_col and matched_col in df.columns:
        if matched_col in numeric_cols:
            return f"{matched_col} statistics:\n\nAverage: {df[matched_col].mean():,.2f}\nMinimum: {df[matched_col].min():,.2f}\nMaximum: {df[matched_col].max():,.2f}"

        counts = df[matched_col].value_counts().head(10)
        return f"Related information from {matched_col}:\n\n{counts.to_string()}"

    return f"""
I could not find an exact intent, but here is a useful dataset summary:

Total Travelers: {len(df)}
Top Destination: {df['Destination'].value_counts().idxmax()}
Most Used Transport: {df['Transport'].value_counts().idxmax()}
Most Common Accommodation: {df['Accommodation'].value_counts().idxmax()}
Average Hotel Cost: ₹{df['Hotel Cost'].mean():,.0f}
Average Transport Cost: ₹{df['Transport Cost'].mean():,.0f}
Average Total Trip Cost: ₹{df['Total Trip Cost'].mean():,.0f}

Try questions like:
- most visited place
- cheapest trip
- average hotel price
- transport used most
- how many male travelers
- show details for Paris
- average cost for Tokyo
"""


st.markdown("""
<div class="hero">
    <h1>🌍 AI Travel Agent Pro</h1>
    <p>Advanced travel analytics, 3D dashboard, smart budget planner, destination insights, transport analysis and downloadable Excel reports.</p>
</div>
""", unsafe_allow_html=True)

uploaded_file = st.sidebar.file_uploader("Upload Travel CSV File", type=["csv"])

if uploaded_file is None:
    st.sidebar.info("Using default file: travel_data.csv")
    file_path = "travel_data.csv"
else:
    file_path = uploaded_file

try:
    df = load_data(file_path)
except Exception as e:
    st.error(f"Dataset loading error: {e}")
    st.stop()

st.sidebar.header("Smart Filters")

destinations = ["All"] + sorted(df["Destination"].dropna().unique().tolist())
selected_destination = st.sidebar.selectbox("Select Destination", destinations)

genders = ["All"] + sorted(df["Gender"].dropna().unique().tolist())
selected_gender = st.sidebar.selectbox("Select Gender", genders)

transports = ["All"] + sorted(df["Transport"].dropna().unique().tolist())
selected_transport = st.sidebar.selectbox("Select Transport", transports)

budget = st.sidebar.number_input("Total Budget", min_value=100, value=2000, step=100)
trip_type = st.sidebar.selectbox("Trip Type", ["Budget", "Standard", "Luxury"])

filtered_df = df.copy()

if selected_destination != "All":
    filtered_df = filtered_df[filtered_df["Destination"] == selected_destination]

if selected_gender != "All":
    filtered_df = filtered_df[filtered_df["Gender"] == selected_gender]

if selected_transport != "All":
    filtered_df = filtered_df[filtered_df["Transport"] == selected_transport]


if filtered_df.empty:
    st.warning("No data found for the selected filters. Please change filters from the sidebar.")

c1, c2, c3, c4, c5 = st.columns(5)

avg_hotel = filtered_df["Hotel Cost"].mean() if not filtered_df.empty else 0
avg_transport = filtered_df["Transport Cost"].mean() if not filtered_df.empty else 0
avg_duration = filtered_df["Duration"].mean() if not filtered_df.empty else 0

metrics = [
    ("Total Travelers", f"{len(filtered_df):,}"),
    ("Destinations", f"{filtered_df['Destination'].nunique():,}" if not filtered_df.empty else "0"),
    ("Avg Hotel Cost", f"₹{avg_hotel:,.0f}" if pd.notna(avg_hotel) else "₹0"),
    ("Avg Transport", f"₹{avg_transport:,.0f}" if pd.notna(avg_transport) else "₹0"),
    ("Avg Duration", f"{avg_duration:.1f} Days" if pd.notna(avg_duration) else "0 Days")
]

for col, (title, value) in zip([c1, c2, c3, c4, c5], metrics):
    with col:
        st.markdown(f"""
        <div class="metric3d">
            <h3>{title}</h3>
            <h2>{value}</h2>
        </div>
        """, unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)


tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📊 Overview",
    "🌍 Destinations",
    "💰 Budget Planner",
    "🤖 Ask AI",
    "📥 Report"
])

with tab1:
    if filtered_df.empty:
        st.info("No chart data available for the selected filters.")
    else:
        col1, col2 = st.columns(2)

        with col1:
            gender_fig = px.pie(
                filtered_df,
                names="Gender",
                title="Traveler Gender Distribution",
                hole=0.45
            )
            gender_fig.update_layout(template="plotly_dark", height=430)
            st.plotly_chart(gender_fig, use_container_width=True)

        with col2:
            transport_fig = px.pie(
                filtered_df,
                names="Transport",
                title="Transport Usage",
                hole=0.45
            )
            transport_fig.update_layout(template="plotly_dark", height=430)
            st.plotly_chart(transport_fig, use_container_width=True)

        age_fig = px.histogram(
            filtered_df,
            x="Age",
            nbins=20,
            title="Traveler Age Distribution",
            color="Gender"
        )
        age_fig.update_layout(template="plotly_dark", height=430)
        st.plotly_chart(age_fig, use_container_width=True)

with tab2:
    if filtered_df.empty:
        st.info("No destination data available for the selected filters.")
    else:
        top_dest = filtered_df["Destination"].value_counts().head(12).reset_index()
        top_dest.columns = ["Destination", "Trips"]

        fig_3d = go.Figure(data=[
            go.Bar(
                x=top_dest["Destination"],
                y=top_dest["Trips"],
                marker=dict(
                    color=top_dest["Trips"],
                    colorscale="Viridis",
                    line=dict(color="white", width=1)
                )
            )
        ])

        fig_3d.update_layout(
            title="3D Style Top Destination Ranking",
            template="plotly_dark",
            height=520,
            xaxis_title="Destination",
            yaxis_title="Number of Trips"
        )

        st.plotly_chart(fig_3d, use_container_width=True)

        scatter = px.scatter(
            filtered_df,
            x="Hotel Cost",
            y="Transport Cost",
            size="Total Trip Cost",
            color="Destination",
            hover_data=["Accommodation", "Transport", "Duration"],
            title="Destination Cost Comparison"
        )
        scatter.update_layout(template="plotly_dark", height=520)
        st.plotly_chart(scatter, use_container_width=True)

with tab3:
    if filtered_df.empty:
        st.info("Budget planner cannot show trips because no data matches the selected filters.")
    else:
        st.subheader("Smart Budget Allocation")

        if trip_type == "Budget":
            hotel_ratio, transport_ratio, activity_ratio, emergency_ratio = 0.35, 0.40, 0.15, 0.10
        elif trip_type == "Luxury":
            hotel_ratio, transport_ratio, activity_ratio, emergency_ratio = 0.50, 0.25, 0.15, 0.10
        else:
            hotel_ratio, transport_ratio, activity_ratio, emergency_ratio = 0.40, 0.35, 0.15, 0.10

        budget_df = pd.DataFrame({
            "Category": ["Hotel", "Transport", "Activities", "Emergency"],
            "Amount": [
                budget * hotel_ratio,
                budget * transport_ratio,
                budget * activity_ratio,
                budget * emergency_ratio
            ]
        })

        col1, col2 = st.columns(2)

        with col1:
            bfig = px.pie(
                budget_df,
                names="Category",
                values="Amount",
                title="Budget Distribution",
                hole=0.45
            )
            bfig.update_layout(template="plotly_dark", height=430)
            st.plotly_chart(bfig, use_container_width=True)

        with col2:
            bar = px.bar(
                budget_df,
                x="Category",
                y="Amount",
                text="Amount",
                title="Budget Category Breakdown"
            )
            bar.update_traces(texttemplate="₹%{text:,.0f}", textposition="outside")
            bar.update_layout(template="plotly_dark", height=430)
            st.plotly_chart(bar, use_container_width=True)

        st.subheader("Recommended Trips Within Your Budget")

        affordable = filtered_df[filtered_df["Total Trip Cost"] <= budget].sort_values("Total Trip Cost")

        if affordable.empty:
            st.warning("No trips found within this budget. Try increasing your budget.")
        else:
            st.dataframe(
                affordable[[
                    "Destination", "Accommodation", "Hotel Cost",
                    "Transport", "Transport Cost", "Total Trip Cost", "Duration"
                ]].head(20),
                use_container_width=True,
                hide_index=True
            )
with tab4:
    st.subheader("Ask Questions From Your Travel Dataset")

    question = st.text_input(
        "Ask anything about your dataset",
        placeholder="Example: most visited place, cheapest trip, average hotel price, transport used most, show Paris details..."
    )

    col_a, col_b, col_c = st.columns(3)

    with col_a:
        if st.button("Most Visited Places", use_container_width=True):
            question = "most visited places"

    with col_b:
        if st.button("Cheapest Trips", use_container_width=True):
            question = "cheapest trips"

    with col_c:
        if st.button("Dataset Summary", use_container_width=True):
            question = "summary"

    if question:
        answer = answer_question(filtered_df, question)

        st.markdown("### Smart Answer")
        st.code(answer, language="text")

    st.info(
        "You do not need exact questions. Similar words work: "
        "place = destination, stay = hotel, vehicle = transport, cheap = low budget, costly = expensive."
    )

with tab5:
    st.subheader("Download Professional Excel Report")

    if filtered_df.empty:
        st.info("No report can be generated because no data matches the selected filters.")
    else:
        report = make_excel_report(df, filtered_df, budget, trip_type)

        st.download_button(
            "Download Excel Report",
            data=report,
            file_name="AI_Travel_Agent_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )

        st.subheader("Filtered Data Preview")
        st.dataframe(filtered_df, use_container_width=True, hide_index=True)