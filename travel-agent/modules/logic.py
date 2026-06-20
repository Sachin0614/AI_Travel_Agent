import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from datetime import datetime

logger = logging.getLogger(__name__)


def calculate_budget_allocation(dataframe, total_budget, trip_type="standard"):
    """
    Calculate smart budget allocation with ML-based hotel recommendation
    
    Args:
        dataframe: Travel data
        total_budget: Total travel budget in INR
        trip_type: Type of trip (standard, luxury, budget)
        
    Returns:
        Tuple of (affordable_hotels, travel_mode, shopping_allowance, best_hotel)
    """
    try:
        # Validate budget
        if not isinstance(total_budget, (int, float)) or total_budget <= 0:
            logger.error(f"Invalid budget: {total_budget}")
            return pd.DataFrame(), "Unknown", 0, "Invalid Budget"
        
        # Adjust allocation based on trip type
        allocation_map = {
            "budget": {"hotel": 0.35, "travel": 0.40, "shopping": 0.15, "misc": 0.10},
            "standard": {"hotel": 0.40, "travel": 0.40, "shopping": 0.15, "misc": 0.05},
            "luxury": {"hotel": 0.50, "travel": 0.30, "shopping": 0.15, "misc": 0.05},
        }
        
        allocation = allocation_map.get(trip_type, allocation_map["standard"])
        
        hotel_budget_limit = total_budget * allocation["hotel"]
        travel_budget_limit = total_budget * allocation["travel"]
        shopping_budget_limit = total_budget * allocation["shopping"]
        
        # Filter affordable hotels
        if 'Cleaned_Price' in dataframe.columns:
            affordable_hotels = dataframe[dataframe['Cleaned_Price'] <= hotel_budget_limit]
        else:
            logger.warning("Cleaned_Price column not found")
            affordable_hotels = pd.DataFrame()
        
        # ML-Based Best Hotel Match using Cosine Similarity
        best_hotel_name = "None (Increase budget)"
        
        if not affordable_hotels.empty and 'Cleaned_Price' in affordable_hotels.columns:
            try:
                prices = affordable_hotels['Cleaned_Price'].values.reshape(-1, 1)
                target = np.array([[hotel_budget_limit]])
                similarities = cosine_similarity(prices, target)
                best_match_idx = np.argmax(similarities)
                best_hotel = affordable_hotels.iloc[best_match_idx]
                best_hotel_name = best_hotel['Hotel_Name'] if 'Hotel_Name' in best_hotel else "Unknown Hotel"
            except Exception as e:
                logger.warning(f"Error in hotel matching: {e}")
        
        # Determine travel mode based on budget
        if travel_budget_limit >= 20000:
            recommended_travel_mode = "✈️ Flight (Recommended)"
        elif travel_budget_limit >= 10000:
            recommended_travel_mode = "🚄 AC Sleeper Train/Premium Bus"
        elif travel_budget_limit >= 5000:
            recommended_travel_mode = "🚌 AC Bus / Standard Train"
        else:
            recommended_travel_mode = "🚌 Budget Bus / Non-AC Train"
        
        logger.info(f"Budget allocation calculated: {travel_budget_limit}, Mode: {recommended_travel_mode}")
        
        return affordable_hotels, recommended_travel_mode, shopping_budget_limit, best_hotel_name
        
    except Exception as e:
        logger.error(f"Error in budget allocation: {e}")
        return pd.DataFrame(), "Unknown", 0, "Error Processing"


def generate_excel_report(males, females, best_hotel, budget, travel_mode, shopping_allowance):
    """
    Generate Excel report with comprehensive itinerary
    
    Returns:
        Excel file as bytes
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Travel Itinerary"
        
        # Define styles
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="004E89")
        border = None
        
        # Title
        ws['A1'] = "🌍 PROFESSIONAL TRAVEL ITINERARY"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Report Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:D2')
        
        # Trip Details Section
        ws['A4'] = "TRIP DETAILS"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        
        trip_details = [
            ("Total Budget", f"₹{budget:,.2f}"),
            ("Recommended Hotel", best_hotel),
            ("Travel Mode", travel_mode),
            ("Shopping Budget", f"₹{shopping_allowance:,.2f}"),
        ]
        
        row = 5
        for label, value in trip_details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Demographics Section
        ws['A11'] = "TRAVELER DEMOGRAPHICS"
        ws['A11'].font = header_font
        ws['A11'].fill = header_fill
        
        ws['A12'] = "Gender"
        ws['B12'] = "Count"
        ws['A12'].font = header_font
        ws['B12'].font = header_font
        ws['A12'].fill = PatternFill(start_color="E8E9F3", end_color="E8E9F3", fill_type="solid")
        ws['B12'].fill = PatternFill(start_color="E8E9F3", end_color="E8E9F3", fill_type="solid")
        
        ws['A13'] = "Male"
        ws['B13'] = males
        ws['A14'] = "Female"
        ws['B14'] = females
        ws['A15'] = "Total"
        ws['B15'] = males + females
        ws['A15'].font = Font(bold=True)
        ws['B15'].font = Font(bold=True)
        
        # Tips Section
        ws['A18'] = "💡 PRO TIPS FOR YOUR TRIP"
        ws['A18'].font = header_font
        ws['A18'].fill = header_fill
        ws.merge_cells('A18:B18')
        
        tips = [
            "1. Book accommodation at least 2-3 weeks in advance for better rates",
            "2. Travel during off-season months for cost savings and fewer crowds",
            "3. Set aside 10% of your budget as emergency contingency",
            "4. Use travel comparison apps to find the best flight and hotel deals",
            "5. Consider travel insurance for protection against cancellations",
        ]
        
        row = 19
        for tip in tips:
            ws[f'A{row}'] = tip
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        
        # Save to bytes
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        logger.info("Excel report generated successfully")
        return file_stream.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return b""