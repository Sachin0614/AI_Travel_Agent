import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import logging
import re
from datetime import datetime
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
from fpdf import FPDF

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# UTILITY FUNCTIONS
# ============================================================================

def clean_currency_and_convert(value):
    """Clean and convert currency values"""
    try:
        if pd.isna(value):
            return 0.0
        
        string_value = str(value).strip()
        for symbol in ['$', 'USD', '€', '₹', '£', ',', ' ']:
            string_value = string_value.replace(symbol, '')
        
        string_value = string_value.strip()
        if string_value:
            return float(string_value)
        else:
            return 0.0
    except (ValueError, TypeError):
        return 0.0


@st.cache_data
def load_and_prepare_dataset(file_path):
    """Load and prepare travel dataset"""
    try:
        dataframe = pd.read_csv(file_path)
        logger.info(f"Loaded dataset with shape: {dataframe.shape}")
        
        # Standardize column names
        column_mapping = {
            'Traveler gender': 'Gender',
            'Traveler Gender': 'Gender',
            'Accommodation cost': 'Hotel_Price',
            'Accommodation type': 'Hotel_Name',
            'Traveler name': 'Name',
            'Traveler age': 'Age',
        }
        
        dataframe.rename(columns=column_mapping, inplace=True)
        
        # Clean the price column
        if 'Hotel_Price' in dataframe.columns:
            dataframe['Cleaned_Price'] = dataframe['Hotel_Price'].apply(clean_currency_and_convert)
        else:
            dataframe['Cleaned_Price'] = 0.0
        
        # Fill missing values
        dataframe.fillna({
            'Gender': 'Unknown',
            'Hotel_Name': 'Not Specified',
            'Cleaned_Price': 0.0,
        }, inplace=True)
        
        logger.info("Dataset preparation completed successfully")
        return dataframe
        
    except FileNotFoundError:
        logger.error(f"File not found: {file_path}")
        return None
    except Exception as e:
        logger.error(f"Error loading dataset: {e}")
        return None


def create_demographic_chart(dataframe):
    """Create demographic pie chart"""
    try:
        if dataframe is None or dataframe.empty:
            return go.Figure().add_annotation(text="No data available")
        
        # Get gender column
        gender_col = None
        for col in ['Gender', 'Traveler gender', 'Traveler Gender']:
            if col in dataframe.columns:
                gender_col = col
                break
        
        if gender_col is None:
            return go.Figure().add_annotation(text="Gender data not available")
        
        # Count genders
        gender_counts = dataframe[gender_col].value_counts()
        
        # Create pie chart
        fig = px.pie(
            values=gender_counts.values,
            names=gender_counts.index,
            title="👥 Traveler Demographics by Gender",
            color_discrete_sequence=['#FF6B35', '#004E89', '#1CAC78'],
            hole=0.3,
        )
        
        fig.update_traces(textposition='inside', textinfo='percent+label')
        fig.update_layout(
            template='plotly_dark',
            font=dict(family="Arial", size=12),
            showlegend=True,
        )
        
        logger.info("Demographic chart created successfully")
        return fig
        
    except Exception as e:
        logger.error(f"Error creating demographic chart: {e}")
        return go.Figure().add_annotation(text="Error creating chart")


def create_destination_chart(dataframe):
    """Create horizontal bar chart showing top destinations"""
    try:
        if dataframe is None or dataframe.empty:
            return go.Figure().add_annotation(text="No data available")
        
        if 'Destination' not in dataframe.columns:
            return go.Figure().add_annotation(text="Destination data not available")
        
        # Get top 10 destinations
        top_destinations = dataframe['Destination'].value_counts().head(10)
        
        # Create horizontal bar chart
        fig = px.bar(
            x=top_destinations.values,
            y=top_destinations.index,
            orientation='h',
            title="📍 Top 10 Most Popular Destinations",
            labels={'x': 'Number of Travelers', 'y': 'Destination'},
            color=top_destinations.values,
            color_continuous_scale='Viridis',
        )
        
        fig.update_layout(
            template='plotly_dark',
            xaxis_title="Number of Travelers",
            yaxis_title="Destination",
            font=dict(family="Arial", size=12),
            showlegend=False,
            height=500,
        )
        
        fig.update_traces(hovertemplate='<b>%{y}</b><br>Travelers: %{x}<extra></extra>')
        
        logger.info("Destination chart created successfully")
        return fig
        
    except Exception as e:
        logger.error(f"Error creating destination chart: {e}")
        return go.Figure().add_annotation(text="Error creating chart")


def create_destination_scatter(dataframe):
    """Create scatter plot for destinations"""
    try:
        if dataframe is None or dataframe.empty:
            return go.Figure()
        
        if 'Destination' not in dataframe.columns:
            return go.Figure()
        
        # Get destination counts
        dest_counts = dataframe['Destination'].value_counts().head(15)
        
        # Create scatter plot
        fig = go.Figure(data=[go.Scatter(
            x=dest_counts.index,
            y=dest_counts.values,
            mode='markers+lines',
            marker=dict(
                size=dest_counts.values / 2,
                color=dest_counts.values,
                colorscale='Viridis',
                showscale=True,
                colorbar=dict(title="Number of<br>Travelers"),
            ),
            line=dict(color='#FF6B35', width=2),
            hovertemplate='<b>%{x}</b><br>Travelers: %{y}<extra></extra>',
        )])
        
        fig.update_layout(
            title="🌍 Global Destination Popularity",
            xaxis_title="Destination",
            yaxis_title="Number of Travelers",
            template='plotly_dark',
            height=400,
            hovermode='closest',
        )
        
        fig.update_xaxes(tickangle=45)
        
        return fig
        
    except Exception as e:
        logger.error(f"Error creating scatter plot: {e}")
        return go.Figure()


def calculate_budget_allocation(dataframe, total_budget, trip_type="standard"):
    """Calculate smart budget allocation"""
    try:
        if not isinstance(total_budget, (int, float)) or total_budget <= 0:
            logger.error(f"Invalid budget: {total_budget}")
            return pd.DataFrame(), "Unknown", 0, "Invalid Budget"
        
        # Budget allocation based on trip type
        allocation_map = {
            "budget": {"hotel": 0.35, "travel": 0.40, "shopping": 0.15, "misc": 0.10},
            "standard": {"hotel": 0.40, "travel": 0.40, "shopping": 0.15, "misc": 0.05},
            "luxury": {"hotel": 0.50, "travel": 0.30, "shopping": 0.15, "misc": 0.05},
        }
        
        allocation = allocation_map.get(trip_type, allocation_map["standard"])
        
        hotel_budget_limit = total_budget * allocation["hotel"]
        travel_budget_limit = total_budget * allocation["travel"]
        shopping_budget_limit = total_budget * allocation["shopping"]
        
        # Filter affordable hotels
        if 'Cleaned_Price' in dataframe.columns:
            affordable_hotels = dataframe[dataframe['Cleaned_Price'] <= hotel_budget_limit]
        else:
            affordable_hotels = pd.DataFrame()
        
        # ML-Based Best Hotel Match
        best_hotel_name = "None (Increase budget)"
        
        if not affordable_hotels.empty and 'Cleaned_Price' in affordable_hotels.columns:
            try:
                prices = affordable_hotels['Cleaned_Price'].values.reshape(-1, 1)
                target = np.array([[hotel_budget_limit]])
                similarities = cosine_similarity(prices, target)
                best_match_idx = np.argmax(similarities)
                best_hotel = affordable_hotels.iloc[best_match_idx]
                best_hotel_name = best_hotel['Hotel_Name'] if 'Hotel_Name' in best_hotel else "Unknown Hotel"
            except Exception as e:
                logger.warning(f"Error in hotel matching: {e}")
        
        # Determine travel mode based on budget
        if travel_budget_limit >= 20000:
            recommended_travel_mode = "✈️ Flight (Recommended)"
        elif travel_budget_limit >= 10000:
            recommended_travel_mode = "🚄 AC Sleeper Train/Premium Bus"
        elif travel_budget_limit >= 5000:
            recommended_travel_mode = "🚌 AC Bus / Standard Train"
        else:
            recommended_travel_mode = "🚌 Budget Bus / Non-AC Train"
        
        logger.info(f"Budget allocation calculated: {travel_budget_limit}, Mode: {recommended_travel_mode}")
        
        return affordable_hotels, recommended_travel_mode, shopping_budget_limit, best_hotel_name
        
    except Exception as e:
        logger.error(f"Error in budget allocation: {e}")
        return pd.DataFrame(), "Unknown", 0, "Error Processing"


def generate_excel_report(males, females, best_hotel, budget, travel_mode, shopping_allowance):
    """Generate Excel report"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Travel Itinerary"
        
        # Define styles
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="004E89")
        
        # Title
        ws['A1'] = "🌍 PROFESSIONAL TRAVEL ITINERARY"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Report Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)
        ws.merge_cells('A2:D2')
        
        # Trip Details Section
        ws['A4'] = "TRIP DETAILS"
        ws['A4'].font = header_font
        ws['A4'].fill = header_fill
        
        trip_details = [
            ("Total Budget", f"₹{budget:,.2f}"),
            ("Recommended Hotel", best_hotel),
            ("Travel Mode", travel_mode),
            ("Shopping Budget", f"₹{shopping_allowance:,.2f}"),
        ]
        
        row = 5
        for label, value in trip_details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Demographics Section
        ws['A11'] = "TRAVELER DEMOGRAPHICS"
        ws['A11'].font = header_font
        ws['A11'].fill = header_fill
        
        ws['A12'] = "Gender"
        ws['B12'] = "Count"
        ws['A12'].font = header_font
        ws['B12'].font = header_font
        
        ws['A13'] = "Male"
        ws['B13'] = males
        ws['A14'] = "Female"
        ws['B14'] = females
        ws['A15'] = "Total"
        ws['B15'] = males + females
        ws['A15'].font = Font(bold=True)
        ws['B15'].font = Font(bold=True)
        
        # Tips Section
        ws['A18'] = "💡 PRO TIPS FOR YOUR TRIP"
        ws['A18'].font = header_font
        ws['A18'].fill = header_fill
        ws.merge_cells('A18:B18')
        
        tips = [
            "1. Book accommodation at least 2-3 weeks in advance for better rates",
            "2. Travel during off-season months for cost savings and fewer crowds",
            "3. Set aside 10% of your budget as emergency contingency",
            "4. Use travel comparison apps to find the best flight and hotel deals",
            "5. Consider travel insurance for protection against cancellations",
        ]
        
        row = 19
        for tip in tips:
            ws[f'A{row}'] = tip
            ws[f'A{row}'].alignment = Alignment(wrap_text=True)
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 30
        
        # Save to bytes
        file_stream = BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        
        logger.info("Excel report generated successfully")
        return file_stream.getvalue()
        
    except Exception as e:
        logger.error(f"Error generating Excel report: {e}")
        return b""


def generate_project_update_pdf(changes, output_path=None):
    """Generate a short PDF summarizing the UI and project updates."""
    try:
        pdf = FPDF(format='A4')
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 18)
        pdf.set_text_color(34, 34, 34)
        pdf.cell(0, 10, "Travel Agent Project Update Summary", ln=True, align='C')
        pdf.ln(5)
        pdf.set_font("Helvetica", "", 12)
        pdf.set_text_color(54, 54, 54)
        pdf.multi_cell(0, 8, "This PDF describes the updates made to the Travel Agent project, including UI improvements and report generation features.")
        pdf.ln(4)

        for idx, item in enumerate(changes, start=1):
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(10, 8, f"{idx}.")
            pdf.set_font("Helvetica", "", 12)
            pdf.multi_cell(0, 8, item)
            pdf.ln(1)

        if output_path:
            pdf.output(output_path)
        return pdf.output(dest='S').encode('latin1')
    except Exception as e:
        logger.error(f"Error generating project summary PDF: {e}")
        return b""


def voice_recorder_component():
    html = """
    <style>
    .voice-card {
        background: rgba(255,255,255,0.08);
        border-radius: 20px;
        border: 1px solid rgba(255,255,255,0.14);
        padding: 1rem;
        color: white;
        font-family: Arial, sans-serif;
    }
    .voice-button {
        background: linear-gradient(135deg, #ff8c42, #ff5c8a);
        border: none;
        color: white;
        font-size: 1rem;
        padding: 0.85rem 1.2rem;
        border-radius: 14px;
        cursor: pointer;
        margin-right: 0.75rem;
    }
    .voice-status {
        margin-top: 0.75rem;
        font-size: 0.95rem;
        color: #d9dce6;
    }
    </style>
    <div class="voice-card">
      <h3>🎙️ Voice Input</h3>
      <p>Click Start, speak your command, then Stop to send the transcript to the dashboard.</p>
      <button class="voice-button" onclick="startRecording()">Start</button>
      <button class="voice-button" onclick="stopRecording()">Stop</button>
      <div id="status" class="voice-status">Status: waiting for input</div>
      <div id="transcript" class="voice-status" style="margin-top:0.5rem;"></div>
    </div>
    <script>
      const SpeechRecognition = window.SpeechRecognition || window.webkitSpeechRecognition;
      const statusElement = document.getElementById('status');
      const transcriptElement = document.getElementById('transcript');
      if (!SpeechRecognition) {
        statusElement.innerText = 'Speech Recognition not supported in this browser.';
      } else {
        const recognition = new SpeechRecognition();
        recognition.interimResults = false;
        recognition.lang = 'en-US';

        recognition.onstart = () => {
          statusElement.innerText = 'Listening... Please speak clearly.';
        };

        recognition.onresult = (event) => {
          const text = event.results[0][0].transcript;
          transcriptElement.innerText = 'Transcript: ' + text;
          window.parent.postMessage({isStreamlitMessage:true, type:'streamlit:setComponentValue', value:text}, '*');
          statusElement.innerText = 'Captured voice input successfully.';
        };

        recognition.onerror = (event) => {
          statusElement.innerText = 'Error: ' + event.error;
        };

        recognition.onend = () => {
          if (statusElement.innerText.startsWith('Listening')) {
            statusElement.innerText = 'Listening stopped.';
          }
        };

        window.startRecording = () => {
          recognition.start();
        };

        window.stopRecording = () => {
          recognition.stop();
        };
      }
    </script>
    """
    return components.html(html, height=260)


# ============================================================================
# MAIN APP
# ============================================================================

def main():
    """Main application"""
    # Page configuration
    st.set_page_config(
        page_title="🚀 Professional Travel Agent",
        page_icon="🚀",
        layout="wide",
        initial_sidebar_state="expanded",
    )

    # Initialize session state defaults for voice control and input persistence
    if "user_input_budget" not in st.session_state:
        st.session_state.user_input_budget = 50000
    if "trip_type" not in st.session_state:
        st.session_state.trip_type = "standard"
    if "trip_duration" not in st.session_state:
        st.session_state.trip_duration = 7
    if "voice_transcript" not in st.session_state:
        st.session_state.voice_transcript = ""

    # Custom CSS
    st.markdown("""
        <style>
        .main-header {
            font-size: 2.8rem;
            color: #FF8C42;
            font-weight: 900;
            letter-spacing: 0.4px;
            margin-bottom: 1rem;
        }
        .stApp {
            background: linear-gradient(180deg, #081f4a 0%, #04102b 100%);
            color: #f7f7fb;
        }
        .stButton>button {
            background: linear-gradient(135deg, #ff8c42, #ff5c8a);
            border: none;
            color: white;
            font-weight: 700;
        }
        .stSidebar .css-1d391kg {
            background: linear-gradient(180deg, #04102b 0%, #092859 100%);
        }
        .hero-banner {
            border-radius: 32px;
            padding: 2rem;
            display: flex;
            justify-content: space-between;
            align-items: center;
            gap: 2rem;
            background: linear-gradient(135deg, rgba(255,140,66,0.14), rgba(11, 32, 75, 0.96));
            border: 1px solid rgba(255,255,255,0.16);
            box-shadow: 0 24px 80px rgba(0, 0, 0, 0.24);
            margin-bottom: 1.5rem;
        }
        .hero-content {
            max-width: 60%;
        }
        .hero-title {
            margin: 0 0 0.75rem 0;
            font-size: 3.4rem;
            line-height: 1.05;
            color: #ffffff;
        }
        .hero-subtitle {
            margin: 0;
            font-size: 1.15rem;
            color: #e2ebff;
            max-width: 95%;
        }
        .hero-image {
            width: 45%;
            min-height: 260px;
            border-radius: 28px;
            background: radial-gradient(circle at top right, rgba(255,140,66,0.35), transparent 28%),
                        radial-gradient(circle at bottom left, rgba(0,194,255,0.25), transparent 18%),
                        linear-gradient(135deg, rgba(255,255,255,0.08), rgba(255,255,255,0.02));
            border: 1px solid rgba(255,255,255,0.14);
            box-shadow: 0 24px 60px rgba(0,0,0,0.20);
            position: relative;
            overflow: hidden;
        }
        .hero-image::before {
            content: "";
            position: absolute;
            top: 20%;
            left: 12%;
            width: 60%;
            height: 60%;
            border-radius: 50%;
            background: rgba(255,255,255,0.12);
            transform: translate(-10%, -10%);
        }
        .hero-image::after {
            content: "";
            position: absolute;
            bottom: 12%;
            right: 10%;
            width: 36%;
            height: 30%;
            background: rgba(255,255,255,0.08);
            border-radius: 18px;
        }
        .metric-card, .dashboard-card, .summary-card {
            background: rgba(255,255,255,0.08);
            border: 1px solid rgba(255,255,255,0.12);
            border-radius: 18px;
            padding: 1.15rem;
            box-shadow: 0 16px 40px rgba(0,0,0,0.22);
        }
        .report-box {
            background: rgba(255,255,255,0.1);
            border-left: 4px solid #ff8c42;
            border-radius: 14px;
            padding: 1rem;
            margin-bottom: 1rem;
        }
        .stDownloadButton>button {
            background: #00c2ff;
            color: #031b3c;
            font-weight: 800;
        }
        .section-header {
            border-radius: 24px;
            background: rgba(255,255,255,0.08);
            padding: 1.5rem;
            margin-bottom: 1.5rem;
            border: 1px solid rgba(255,255,255,0.14);
        }
        .info-box {
            background: rgba(255,255,255,0.06);
            border-radius: 18px;
            border: 1px solid rgba(255,255,255,0.12);
            padding: 1.2rem;
            margin-bottom: 1.5rem;
            color: #f3f4f8;
        }
        .info-box h3 {
            margin-top: 0;
            color: #fff;
        }
        .info-box p {
            margin-bottom: 0;
            line-height: 1.6;
            color: #d9dce6;
        }
        .chart-card {
            border-radius: 20px;
            overflow: hidden;
            padding: 1rem;
            background: rgba(255,255,255,0.05);
            border: 1px solid rgba(255,255,255,0.12);
        }
        .download-row {
            margin-top: 1.2rem;
        }
        </style>
    """, unsafe_allow_html=True)
    
    # ===== SIDEBAR =====
    st.sidebar.title("⚙️ Control Panel")
    st.sidebar.write("Customize your travel parameters below:")
    
    # Budget input
    user_input_budget = st.sidebar.number_input(
        "💰 Total Travel Budget (in INR):",
        min_value=5000,
        value=st.session_state.user_input_budget,
        step=5000,
        help="Enter your total travel budget in Indian Rupees",
        key="user_input_budget",
    )
    
    # Trip type selector
    trip_type = st.sidebar.selectbox(
        "🏷️ Trip Type:",
        ["standard", "budget", "luxury"],
        index=["standard", "budget", "luxury"].index(st.session_state.trip_type),
        help="Choose your trip style to optimize budget allocation",
        key="trip_type",
    )
    
    # Duration input
    duration = st.sidebar.slider(
        "📅 Trip Duration (days):",
        min_value=1,
        max_value=30,
        value=st.session_state.trip_duration,
        key="trip_duration",
    )
    
    # Generate button
    generate_button = st.sidebar.button(
        "🚀 Generate Smart Itinerary",
        type="primary",
        use_container_width=True,
    )
    
    st.sidebar.markdown("---")
    st.sidebar.info(
        "💡 **Tip:** This app analyzes travel data to provide personalized recommendations based on your budget and preferences."
    )
    
    # ===== MAIN CONTENT =====
    st.markdown(
        '<div class="hero-banner">'
        '  <div class="hero-content">'
        '    <h1 class="hero-title">🚀 Professional Travel Intelligence Dashboard</h1>'
        '    <p class="hero-subtitle">Use data-driven insights and voice commands to plan your next trip. Get recommendations, export reports, and explore destination trends in one polished dashboard.</p>'
        '  </div>'
        '  <div class="hero-image"></div>'
        '</div>'
        '</div>',
        unsafe_allow_html=True,
    )
    st.markdown("---")
    
    # Voice command recorder
    voice_recorder_component()

    st.markdown(
        """
        <div class="info-box">
            <h3>🎧 Voice Command Examples</h3>
            <p>Use one of these phrases with the recorder or paste the transcript below:</p>
            <ul>
                <li><strong>Set budget 70000</strong></li>
                <li><strong>Luxury trip for 5 days</strong></li>
                <li><strong>Standard trip for 7 days</strong></li>
            </ul>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.text_area(
        "Voice transcript",
        value=st.session_state.voice_transcript,
        height=90,
        key="voice_transcript",
        help="If needed, paste or edit the transcript here for voice command parsing.",
    )

    voice_input = st.session_state.voice_transcript if isinstance(st.session_state.voice_transcript, str) else ""
    if voice_input:
        st.success("🎤 Voice command received")
        parsed = voice_input.lower()
        if 'luxury' in parsed:
            st.session_state.trip_type = 'luxury'
        elif 'standard' in parsed:
            st.session_state.trip_type = 'standard'
        elif 'budget' in parsed:
            st.session_state.trip_type = 'budget'

        budget_match = re.search(r'([0-9]{4,7})', parsed.replace(',', ''))
        if budget_match:
            st.session_state.user_input_budget = int(budget_match.group(1))

        duration_match = re.search(r'([0-9]{1,2})\s*(day|days)', parsed)
        if duration_match:
            st.session_state.trip_duration = int(duration_match.group(1))

        if 'destination' in parsed:
            st.info("Tip: Voice commands affecting destination filtering are coming soon.")

    # Re-sync local values with session state after voice input
    user_input_budget = st.session_state.user_input_budget
    trip_type = st.session_state.trip_type
    duration = st.session_state.trip_duration

    # Load dataset
    dataset_file_path = "travel_data.csv"
    travel_dataframe = load_and_prepare_dataset(dataset_file_path)
    
    if travel_dataframe is None:
        st.error("❌ Error: 'travel_data.csv' file not found. Please ensure the file is in the app directory.")
        return
    
    # ===== STATISTICS =====
    st.subheader("📈 Real-Time Global Statistics")
    
    with st.container():
        stat1, stat2, stat3, stat4 = st.columns(4)
        with stat1:
            st.markdown('<div class="metric-card">', unsafe_allow_html=True)
            st.metric(label="👥 Total Travelers", value=f"{len(travel_dataframe):,}", delta="Active Records")
            st.markdown('</div>', unsafe_allow_html=True)
        with stat2:
            top_country = travel_dataframe['Destination'].value_counts().index[0] if 'Destination' in travel_dataframe.columns else "N/A"
            st.markdown('<div class="metric-card">', unsafe_allow_html=True)
            st.metric(label="🌍 Most Popular", value=top_country)
            st.markdown('</div>', unsafe_allow_html=True)
        with stat3:
            avg_cost = travel_dataframe['Cleaned_Price'].mean()
            st.markdown('<div class="metric-card">', unsafe_allow_html=True)
            st.metric(label="💵 Average Cost", value=f"₹{avg_cost:,.0f}")
            st.markdown('</div>', unsafe_allow_html=True)
        with stat4:
            unique_dests = travel_dataframe['Destination'].nunique() if 'Destination' in travel_dataframe.columns else 0
            st.markdown('<div class="metric-card">', unsafe_allow_html=True)
            st.metric(label="📍 Destinations", value=f"{unique_dests}", delta="Covered")
            st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    
    # ===== VISUALIZATIONS =====
    st.subheader("📊 Demographic & Destination Insights")
    
    viz_col1, viz_col2 = st.columns(2)
    
    with viz_col1:
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.plotly_chart(
            create_demographic_chart(travel_dataframe),
            use_container_width=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)
    
    with viz_col2:
        st.markdown('<div class="chart-card">', unsafe_allow_html=True)
        st.plotly_chart(
            create_destination_chart(travel_dataframe),
            use_container_width=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)
    
    # Destination Scatter Map - REPLACES FOLIUM MAP
    st.subheader("🌍 Global Destination Popularity Chart")
    st.markdown('<div class="chart-card">', unsafe_allow_html=True)
    st.plotly_chart(
        create_destination_scatter(travel_dataframe),
        use_container_width=True,
    )
    st.markdown('</div>', unsafe_allow_html=True)
    
    st.markdown("---")
    
    top_destins = travel_dataframe['Destination'].value_counts().head(6) if 'Destination' in travel_dataframe.columns else pd.Series()
    if not top_destins.empty:
        st.subheader("✨ Top Destinations Snapshot")
        dest_col1, dest_col2 = st.columns([2, 1])
        with dest_col1:
            st.bar_chart(top_destins.rename_axis('Destination').reset_index(name='Count').set_index('Destination'))
        with dest_col2:
            st.markdown('<div class="report-box">', unsafe_allow_html=True)
            st.markdown('### 📌 Quick insights')
            for dest, count in top_destins.items():
                st.write(f"**{dest}** — {count:,} trips")
            st.markdown('</div>', unsafe_allow_html=True)
        st.markdown("---")
    
    # ===== SMART RECOMMENDATIONS =====
    if generate_button:
        st.subheader("💡 Your Personalized Travel Itinerary")
        
        with st.spinner("🔄 Analyzing travel data and generating recommendations..."):
            # Calculate budget allocation
            affordable_hotels, recommended_travel, shopping_allowance, best_hotel = calculate_budget_allocation(
                travel_dataframe,
                user_input_budget,
                trip_type,
            )
            
            # Display recommendations
            rec_col1, rec_col2 = st.columns([1, 2])
            
            with rec_col1:
                st.info(f"✈️ **Recommended Travel Mode:**\n{recommended_travel}")
                st.success(f"🛍️ **Shopping & Misc Fund:**\n₹{shopping_allowance:,.0f}")
                st.warning(f"🏨 **Best Hotel Match:**\n{best_hotel}")
            
            with rec_col2:
                if not affordable_hotels.empty:
                    display_df = affordable_hotels[['Hotel_Name', 'Destination', 'Cleaned_Price']].drop_duplicates().head(10)
                    display_df = display_df.rename(columns={
                        'Hotel_Name': 'Accommodation',
                        'Cleaned_Price': 'Price (₹)',
                    })
                    st.dataframe(display_df, use_container_width=True, hide_index=True)
                else:
                    st.warning("⚠️ No affordable hotels found for this budget. Try increasing your budget.")
            
            # Budget breakdown
            st.markdown("### 💰 Budget Allocation Breakdown")
            
            allocation_col1, allocation_col2 = st.columns(2)
            
            with allocation_col1:
                hotel_budget = user_input_budget * 0.40
                transport_budget = user_input_budget * 0.40
                activities_budget = user_input_budget * 0.15
                misc_budget = user_input_budget * 0.05
                
                breakdown_data = {
                    'Category': ['Accommodation', 'Transport', 'Activities', 'Miscellaneous'],
                    'Amount': [hotel_budget, transport_budget, activities_budget, misc_budget],
                }
                
                fig = px.pie(
                    values=breakdown_data['Amount'],
                    names=breakdown_data['Category'],
                    title="Budget Distribution",
                )
                st.plotly_chart(fig, use_container_width=True)
            
            with allocation_col2:
                st.write(f"🏨 **Accommodation:** ₹{hotel_budget:,.0f}")
                st.write(f"✈️ **Transport:** ₹{transport_budget:,.0f}")
                st.write(f"🎭 **Activities:** ₹{activities_budget:,.0f}")
                st.write(f"🎁 **Miscellaneous:** ₹{misc_budget:,.0f}")
            
            # Export Excel report
            males = len(travel_dataframe[travel_dataframe['Gender'] == 'Male']) if 'Gender' in travel_dataframe.columns else 0
            females = len(travel_dataframe[travel_dataframe['Gender'] == 'Female']) if 'Gender' in travel_dataframe.columns else 0
            
            excel_file = generate_excel_report(
                males,
                females,
                best_hotel,
                user_input_budget,
                recommended_travel,
                shopping_allowance,
            )

            summary_changes = [
                "Designed a polished hero banner, refined dark theme, and premium dashboard card layout.",
                "Built browser-based voice command support with speech recognition and transcript parsing.",
                "Added data-based budget recommendations, travel mode guidance, and affordable hotel matching.",
                "Enabled exportable Excel itinerary reports and a downloadable project summary PDF.",
            ]
            summary_pdf = generate_project_update_pdf(summary_changes, output_path="project_update_summary.pdf")
            
            st.markdown('<div class="download-row">', unsafe_allow_html=True)
            st.download_button(
                label="📥 Download Itinerary Report (Excel)",
                data=excel_file,
                file_name=f"Travel_Itinerary_{user_input_budget}_{trip_type}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.download_button(
                label="📝 Download Project Summary PDF",
                data=summary_pdf,
                file_name="project_update_summary.pdf",
                mime="application/pdf",
                use_container_width=True,
            )
            st.markdown('</div>', unsafe_allow_html=True)


if __name__ == "__main__":
    main()